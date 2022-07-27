import requests
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Border, Side, Font, Fill, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import json

print('--------------'+'Codeforces contest and upsolving extracting'+'--------------')
contest_id = input('Input contest id : ')
url = 'https://codeforces.com/contest/' + contest_id

# finding contest name
html_data = BeautifulSoup(requests.get(url).content, 'html.parser')
html_data = html_data.find('div', id = 'body')
html_data = html_data.find('div', class_ = 'roundbox sidebox')
contest_name = html_data.find('th', class_ = 'left').text.strip()
print(contest_name)

# conntecting to exel file and getting the handles
xl_file = openpyxl.load_workbook('apitest.xlsx')
xl_file2 = openpyxl.load_workbook('test.xlsx')
handles_sheet = xl_file['handles']
handles = [] # for storing the handles from exel file
col_pos = 3 # 3 for codeforces handles
row_pos = 2
for i in handles_sheet.rows:
   handles.append(handles_sheet.cell(row_pos, col_pos).value)
   row_pos += 1

# collecting the data
data = [] # for storing the solve and upsolve count in the order handles

for i in range(0, len(handles) - 1, 1):
   handle_name = handles[i]
   print('Analyzing for : ' + handle_name)
   api_link = 'https://codeforces.com/api/contest.status?contestId=' + contest_id + '&handle="' +handle_name+ '"&from=1&count=50'
   # print(api_link)
   
   # checking if its a json file if not sending requests multiples times
   
   totReqSent = 0
   while (1):
      participant_data = requests.get(api_link)
      check_data_if_json = participant_data.text.strip()
      totReqSent += 1
      if (totReqSent == 5):
         print("Codeforces Not Responding")
         exit()
      try:
         valid_json_file = json.loads(check_data_if_json)
         break
      except ValueError as e:
         continue
      
   
   participant_data = participant_data.json()
   if (participant_data['status'] == 'FAILED'):
      data.append(['Invalid id', 'Invalid id'])
      continue

   contest_time_solve = 0
   upsolve = 0
   p_l1 = []
   p_l2 = []

   data_len = len(participant_data['result'])
   for j in range(0, data_len, 1):
      verdict = participant_data['result'][j]['verdict']
      is_contestant = participant_data['result'][j]['author']['participantType']
      problem_indx = participant_data['result'][j]['problem']['index']

      if (verdict == 'OK'):
         if (is_contestant == 'CONTESTANT' or is_contestant == 'OUT_OF_COMPETITION'):
            if (problem_indx in p_l1): continue
            else: p_l1.append(problem_indx)
         else:
            if (problem_indx in p_l2): continue
            else: p_l2.append(problem_indx)
            
   
   contest_time_solve = len(p_l1)
   upsolve = len(p_l2)

   data.append([contest_time_solve, upsolve]) # stored data

print(data)

print(xl_file.sheetnames)
# sending and saving data in exel file
contest_records = xl_file['Contest  records']
mx_col = 5
mx_row = 3
while 1:
   if (contest_records.cell(2, mx_col).value == None):
      break
   else: mx_col += 1

# upper and lower cell writing
center_alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
contest_records.merge_cells(start_row = 1, start_column = mx_col, end_row = 1, end_column = mx_col + 1)
contest_records.cell(1, mx_col).value = contest_name
contest_records.cell(1, mx_col).font = Font(size=11, underline='single', color='2DA5E7', bold=True, italic=False)
contest_records.cell(1, mx_col).alignment = center_alignment
contest_records.cell(1, mx_col).hyperlink = url

contest_records.cell(2, mx_col).value = 'Solved(contest)'
contest_records.cell(2, mx_col + 1).value = 'Upsolved'
contest_records.cell(2, mx_col).font = Font(size=11, color='E7492D', bold=True, italic=True)
contest_records.cell(2, mx_col + 1).font = Font(size=11, color='E7492D', bold=True, italic=True)
contest_records.cell(2, mx_col).alignment = center_alignment
contest_records.cell(2, mx_col + 1).alignment = center_alignment
contest_records.column_dimensions[get_column_letter(mx_col)].width = max(17, len(contest_name)/2)
contest_records.column_dimensions[get_column_letter(mx_col + 1)].width = max(17, len(contest_name)/2)

# storing data in exel file
for i in data:
   contest_records.cell(mx_row, mx_col).value = i[0]
   contest_records.cell(mx_row, mx_col + 1).value = i[1]
   solved = 0
   upsolved = 0
   if (i[0] != 'Invalid id'): solved = i[0]
   if (i[1] != 'Invalid id'): upsolved = i[1]
   contest_records.cell(mx_row, 3).alignment = center_alignment
   contest_records.cell(mx_row, mx_col).alignment = center_alignment
   contest_records.cell(mx_row, mx_col + 1).alignment = center_alignment
   if (i[0] != 'Invalid id'):
      total_points = contest_records.cell(mx_row, 4).value + 0.0
      total_points += (i[0] * 2 + float(i[1]) * 1)
      contest_records.cell(mx_row, 4).value = total_points

      total_solved = contest_records.cell(mx_row, 3).value + 0
      total_solved += solved + upsolved
      contest_records.cell(mx_row, 3).value = total_solved
   
   mx_row += 1


xl_file.save('apitest.xlsx')
# xl_file2 = xl_file
# xl_file2.save('test.xlsx')