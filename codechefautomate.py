import requests
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Border, Side, Font, Fill, Alignment, PatternFill
from openpyxl.utils import get_column_letter

print('--------------'+'Codeforces contest and upsolving extracting'+'--------------')
contest_id = input('Input contest id : ')
url_problem_code = 'https://www.codechef.com/api/contests/' + contest_id + '?'
url = 'https://competitive-coding-api.herokuapp.com/api/codechef/'

# getting problem codes, contest_name of the contest
problem_codes = [] # for holding contest codes
contest_name = '' # for holding contest name
json_data = {} # for holding various json file
try:
   json_data = requests.get(url_problem_code).json()
except:
   print("Please check your connection & try again")
   exit()

contest_name = json_data['name']
problem_set = json_data['problems']
for code in problem_set.keys():
   problem_codes.append(code)

print(problem_codes)


# conntecting to exel file and getting the handles
xl_file2 = openpyxl.load_workbook('apitest.xlsx')
xl_file = openpyxl.load_workbook('test.xlsx')
handles_sheet = xl_file2['handles']
handles = [] # for storing the handles from exel file
col_pos = 2 # 2 for codechef handles
row_pos = 2
for i in handles_sheet.rows:
   handles.append(handles_sheet.cell(row_pos, col_pos).value)
   row_pos += 1


# collecting the data
data = [] # for storing the solve and upsolve count in the order handles

for i in range(0, len(handles), 1):
   handle_name = handles[i]
   if (handle_name == None):
      data.append(['Invalid id', 'Invalid id'])
      continue

   print('Analyzing for : ' + handle_name)
   api_link = url + handle_name
   participant_data = requests.get(api_link).json()

   if (participant_data['status'] == 'Failed'):
      data.append(['Invalid id', 'Invalid id'])
      continue
   
   contest_time_solve = 0
   upsolve = 0

   if (contest_id in participant_data['fully_solved']):
      contest_time_solve = len(participant_data['fully_solved'][contest_id])
   
   if ('Practice' not in participant_data['fully_solved']):
      data.append([contest_time_solve, upsolve])
      continue

   for code in problem_codes:
      for _code in participant_data['fully_solved']['Practice']:
         if (code == _code): upsolve += 1
   
   data.append([contest_time_solve, upsolve])
#----------------------collcting data end---------------------------

print(data)



# sending and saving data in exel file
contest_records = xl_file['Contest  records']
mx_col = 5
mx_row = 3
while 1:
   if (contest_records.cell(2, mx_col).value == None):
      break
   else: mx_col += 1

# upper and lower cell writing
center_alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
contest_records.merge_cells(start_row = 1, start_column = mx_col, end_row = 1, end_column = mx_col + 1)
contest_records.cell(1, mx_col).value = contest_name
contest_records.cell(1, mx_col).font = Font(size=11, underline='single', color='2DA5E7', bold=True, italic=False)
contest_records.cell(1, mx_col).alignment = center_alignment
contest_records.cell(1, mx_col).hyperlink = url

contest_records.cell(2, mx_col).value = 'Solved(contest)'
contest_records.cell(2, mx_col + 1).value = 'Upsolved'
contest_records.cell(2, mx_col).font = Font(size=11, color='E7492D', bold=True, italic=True)
contest_records.cell(2, mx_col + 1).font = Font(size=11, color='E7492D', bold=True, italic=True)
contest_records.cell(2, mx_col).alignment = center_alignment
contest_records.cell(2, mx_col + 1).alignment = center_alignment
contest_records.column_dimensions[get_column_letter(mx_col)].width = min(17, len(contest_name)/2)
contest_records.column_dimensions[get_column_letter(mx_col + 1)].width = min(17, len(contest_name)/2)

# storing data in exel file
for i in data:
   contest_records.cell(mx_row, mx_col).value = i[0]
   contest_records.cell(mx_row, mx_col + 1).value = i[1]
   solved = 0
   upsolved = 0
   if (i[0] != 'Invalid id'): solved = i[0]
   if (i[1] != 'Invalid id'): upsolved = i[1]
   contest_records.cell(mx_row, 3).alignment = center_alignment
   contest_records.cell(mx_row, mx_col).alignment = center_alignment
   contest_records.cell(mx_row, mx_col + 1).alignment = center_alignment
   if (i[0] != 'Invalid id'):
      total_points = contest_records.cell(mx_row, 4).value + 0.0
      total_points += (i[0] * 2 + float(i[1]) * 1)
      contest_records.cell(mx_row, 4).value = total_points

      total_solved = contest_records.cell(mx_row, 3).value + 0
      total_solved += solved + upsolved
      contest_records.cell(mx_row, 3).value = total_solved
   
   mx_row += 1


xl_file.save('test.xlsx')
# xl_file2 = xl_file
# xl_file2.save('test.xlsx')


