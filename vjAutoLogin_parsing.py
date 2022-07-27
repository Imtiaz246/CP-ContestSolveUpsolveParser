import requests
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Border, Side, Font, Fill, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from pprint import pprint

print("--------------Virtual Judge contest and upsolving extracting--------------")
contest_id = input('Input contest id : ')
url = 'https://vjudge.net/contest/' + str(contest_id)
html_data = BeautifulSoup(requests.get(url).content, 'html.parser')

# findding start time and end time and contest name of the contest
html_data = html_data.find('div', class_ = 'container')
startTime = html_data.find('div', class_ = 'col-xs-3 text-xs-left')
startTime = int(startTime.find('span', class_ = 'timestamp').text)
endTime = html_data.find('div', class_ = 'col-xs-3 text-xs-right')
endTime = int(endTime.find('span', class_ = 'timestamp').text)
contest_name = html_data.find('div', class_ = 'col-xs-6 text-xs-center').text.strip()

# conntecting to exel file and getting the handles
xl_file = openpyxl.load_workbook('apitest.xlsx')
xl_file2 = openpyxl.load_workbook('test.xlsx')
handles_sheet = xl_file['handles']
handles = [] # for storing the handles from exel file
col_pos = 6 # 6 for vjudge handles
row_pos = 2
for i in handles_sheet.rows:
   handles.append(handles_sheet.cell(row = row_pos, column = col_pos).value)
   row_pos += 1

print(len(handles))
print(handles)

#finding all the status of the contest
url_for_status1 = 'https://vjudge.net/status/data/?draw=5&start='
url_for_status2 = '&length=20&un=&num=-&res=0&language=&inContest=true&contestId=' + str(contest_id)
start = 0
total_sub = 0

status_data = [] # holding all the status submissons parameter
cnt = 1
while 1:
   print("Requesting status page ", cnt)
   cnt += 1
   status = requests.get(url_for_status1 + str(start) + url_for_status2).json()
   status_data += status['data']
   if (len(status['data']) == 0):
      break
   else:
      total_sub += len(status['data'])
      start += len(status['data'])

team_info = {} # holding the team information (name, solved_in_colntest_list, upsolved_list, cnt_solved, cnt_upsolved)

# finding solved details
for i in range(len(status_data)):
   if (status_data[i]['status'] == 'Accepted' or status_data[i]['status'] == 'Happy New Year!'):
      name = status_data[i]['userName'].lower()
      problem_no = ord(status_data[i]['contestNum'][0]) - ord('A')
      submitTime = status_data[i]['time']

      sol_in_contest = False
      if (submitTime >= startTime and submitTime <= endTime):
         sol_in_contest = True

      found = False
      if (name in team_info):
         if (sol_in_contest):
            team_info[name][0][problem_no] = 1
         else:
            team_info[name][1][problem_no] = 1
         
         found = True

      if (found == False):
         l1 = [None] * 20
         l2 = [None] * 20
         if (sol_in_contest):
            l1[problem_no] = 1
         else:
            l2[problem_no] = 1
         # team_info.append([name, l1, l2, 0, 0])
         team_info[name] = [l1, l2, 0, 0]
         

for i in team_info:
   for j in range(0, 20, 1):
      if (team_info[i][0][j] == 1):
         team_info[i][2] += 1
      else:
         if (team_info[i][1][j] == 1):
            team_info[i][3] += 1
         


print('-------------'+contest_name+'-------------')
# debug
for i in team_info:
   print(i)
   print(team_info[i][2])
   print(team_info[i][3])
   print()


# print(team_info)
# getting row wise data
data = [] # for storing data(solved, upsolved)
for name in handles:
   if (name == None): 
      data.append(['-', '-'])
      continue
   name = str(name).lower()
   if (name in team_info):
      data.append([team_info[name][2], team_info[name][3]])
      # print(name)
   else: data.append(['-', '-'])


# sending and saving data in exel file
contest_records = xl_file['Contest  records']
mx_col = 5
mx_row = 3
while 1:
   if (contest_records.cell(2, mx_col).value == None):
      break
   else: mx_col += 1

# upper and lower cell writing
center_alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
contest_records.merge_cells(start_row = 1, start_column = mx_col, end_row = 1, end_column = mx_col + 1)
contest_records.cell(1, mx_col).value = contest_name
contest_records.cell(1, mx_col).font = Font(size=11, underline='single', color='2DA5E7', bold=True, italic=False)
contest_records.cell(1, mx_col).alignment = center_alignment
contest_records.cell(1, mx_col).hyperlink = url

contest_records.cell(2, mx_col).value = 'Solved(contest)'
contest_records.cell(2, mx_col + 1).value = 'Upsolved'
contest_records.cell(2, mx_col).font = Font(size=11, color='E7492D', bold=True, italic=True)
contest_records.cell(2, mx_col + 1).font = Font(size=11, color='E7492D', bold=True, italic=True)
contest_records.cell(2, mx_col).alignment = center_alignment
contest_records.cell(2, mx_col + 1).alignment = center_alignment
contest_records.column_dimensions[get_column_letter(mx_col)].width = max(17, len(contest_name)/2)
contest_records.column_dimensions[get_column_letter(mx_col + 1)].width = max(17, len(contest_name)/2)

# storing data in exel file
for i in data:
   contest_records.cell(mx_row, mx_col).value = i[0]
   contest_records.cell(mx_row, mx_col + 1).value = i[1]
   solved = 0
   upsolved = 0
   if (i[0] != '-'): solved = i[0]
   if (i[1] != '-'): upsolved = i[1]
   contest_records.cell(mx_row, 3).alignment = center_alignment
   contest_records.cell(mx_row, mx_col).alignment = center_alignment
   contest_records.cell(mx_row, mx_col + 1).alignment = center_alignment
   if (i[0] != '-'):
      total_points = contest_records.cell(mx_row, 4).value + 0.0
      total_points += (i[0] * 2 + float(i[1]) * 1)
      contest_records.cell(mx_row, 4).value = total_points

      total_solved = contest_records.cell(mx_row, 3).value + 0
      total_solved += solved + upsolved
      contest_records.cell(mx_row, 3).value = total_solved
   
   mx_row += 1


xl_file.save('apitest.xlsx')

print(len(team_info))

   




