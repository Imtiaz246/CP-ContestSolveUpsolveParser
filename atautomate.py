import requests
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Border, Side, Font, Fill, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from pprint import pprint
from datetime import datetime

print("--------------Atcoder contest and upsolving extracting--------------")
contest_id = input('Input contest id : ')
url = 'https://atcoder.jp/contests/' + str(contest_id)
html_data = BeautifulSoup(requests.get(url).content, 'html.parser')

# findding start time and end time and contest name of the contest
html_data = html_data.find('div', id = 'main-container')
html_data = html_data.find('div', class_ = 'row')
fullTime = html_data.find_all('time')
# converting time to unix time
start = None
end = None
for time in fullTime:
   if (start == None):
      start = time.text.strip()
   else: end = time.text.strip()

startTime = start.replace('+0900', '')
endTime = end.replace('+0900', '')

startTime = datetime.strptime(startTime, '%Y-%m-%d %H:%M:%S')
endTime = datetime.strptime(endTime, '%Y-%m-%d %H:%M:%S')
startTime = int(startTime.timestamp())
endTime = int(endTime.timestamp())
# print(startTime)
# print(endTime)
contest_name = html_data.find('h1', class_ = 'text-center').text.strip()
print(contest_name)
# ---------------------------------------------------------------------------------

# conntecting to exel file and getting the handles
xl_file = openpyxl.load_workbook('apitest.xlsx')
xl_file2 = openpyxl.load_workbook('test.xlsx')
handles_sheet = xl_file['handles']
handles = [] # for storing the handles from exel file
col_pos = 5 # 5 for atcoder handles
row_pos = 2
for i in handles_sheet.rows:
   handles.append(handles_sheet.cell(row = row_pos, column = col_pos).value)
   row_pos += 1

for i in range(0, len(handles), 1):
   handles[i] = str(handles[i]).lower()
   

print(len(handles))
pprint(handles)
# ---------------------------------------------------------------------------------

# collecting the data
data = [] # for storing the solve and upsolve count in the order handles

cnt = 1
participant_data = [] # for holding participant data in order of handles
for name in handles:
   print('Analyzing data for ' + name)
   if (name == None):
      participant_data.append(None)
   organized_data = [] # for holding table data (submit_time, problem_name)
   page = 1 # for finding all the submission page
   while (1):
      url = 'https://atcoder.jp/contests/' + contest_id + '/submissions?f.Task=&f.LanguageName=&f.Status=AC&f.User='+ name + '&page=' + str(page)
      page += 1 # incrementing page for further data 
      submission_data = BeautifulSoup(requests.get(url).content, 'html.parser')
      submission_data = submission_data.find('div', id = 'main-container')
      submission_data = submission_data.find('div', class_ = 'row')
      submission_data = submission_data.find('div', class_ = 'table-responsive')
      if (submission_data == None):
         break
      submission_data = submission_data.find('tbody')
      submission_data = submission_data.find_all('tr')
      
      # finding the submission details
      for i in range(0, len(submission_data), 1):
         submit_time = submission_data[i].find('time', class_ = 'fixtime-second').text.strip()
         problem_name = submission_data[i].find('a').text.strip()
         submitTime = submit_time.replace('+0900', '')
         submitTime = datetime.strptime(submitTime, '%Y-%m-%d %H:%M:%S')
         submitTime = int(submitTime.timestamp())

         organized_data.append([submitTime, problem_name]) # pushing data in as organized data
   
   participant_data.append(organized_data)

# from participant data extracting solve and upsolve data and assigning to data[] in handle order
for i in range(0, len(participant_data), 1):
   if (participant_data == None):
      data.append(['-', '-'])

   solved = 0
   upsolved = 0
   l1 = []
   l2 = []
   for j in range(0, len(participant_data[i]), 1):
      submit_time = participant_data[i][j][0]
      problem_name = participant_data[i][j][1]
      if (submit_time >= startTime and submit_time <= endTime):
         if (problem_name not in l1):
            l1.append(problem_name)
            solved += 1

   for j in range(0, len(participant_data[i]), 1):
      submit_time = participant_data[i][j][0]
      problem_name = participant_data[i][j][1]
      if (submit_time < startTime or submit_time > endTime):
         if (problem_name not in l1 and problem_name not in l2):
            l2.append(problem_name)
            upsolved += 1

   data.append([solved, upsolved])

   # pprint(data)

   # sending and saving data in exel file
contest_records = xl_file['Contest  records']
mx_col = 5
mx_row = 3
while 1:
   if (contest_records.cell(2, mx_col).value == None):
      break
   else: mx_col += 1

# upper and lower cell writing
center_alignment = Alignment(horizontal='center', vertical='center', text_rotation=0, wrap_text=False, shrink_to_fit=False, indent=0)
contest_records.merge_cells(start_row = 1, start_column = mx_col, end_row = 1, end_column = mx_col + 1)
contest_records.cell(1, mx_col).value = contest_name
contest_records.cell(1, mx_col).font = Font(size=11, underline='single', color='2DA5E7', bold=True, italic=False)
contest_records.cell(1, mx_col).alignment = center_alignment
contest_records.cell(1, mx_col).hyperlink = url

contest_records.cell(2, mx_col).value = 'Solved(contest)'
contest_records.cell(2, mx_col + 1).value = 'Upsolved'
contest_records.cell(2, mx_col).font = Font(size=11, color='E7492D', bold=True, italic=True)
contest_records.cell(2, mx_col + 1).font = Font(size=11, color='E7492D', bold=True, italic=True)
contest_records.cell(2, mx_col).alignment = center_alignment
contest_records.cell(2, mx_col + 1).alignment = center_alignment
contest_records.column_dimensions[get_column_letter(mx_col)].width = max(17, len(contest_name)/2)
contest_records.column_dimensions[get_column_letter(mx_col + 1)].width = max(17, len(contest_name)/2)

# storing data in exel file
for i in data:
   contest_records.cell(mx_row, mx_col).value = i[0]
   contest_records.cell(mx_row, mx_col + 1).value = i[1]
   solved = 0
   upsolved = 0
   if (i[0] != '-'): solved = i[0]
   if (i[1] != '-'): upsolved = i[1]
   contest_records.cell(mx_row, 3).alignment = center_alignment
   contest_records.cell(mx_row, mx_col).alignment = center_alignment
   contest_records.cell(mx_row, mx_col + 1).alignment = center_alignment
   if (i[0] != '-'):
      total_points = contest_records.cell(mx_row, 4).value + 0.0
      total_points += (i[0] * 2 + float(i[1]) * 1)
      contest_records.cell(mx_row, 4).value = total_points

      total_solved = contest_records.cell(mx_row, 3).value + 0
      total_solved += solved + upsolved
      contest_records.cell(mx_row, 3).value = total_solved
   
   mx_row += 1


xl_file.save('apitest.xlsx')

